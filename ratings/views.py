"""
Ratings — public anonymous rating + admin management/dashboard/export.

Public:
  GET  /rate/                       -> list of active facilities
  GET  /rate/<facility_id>/         -> star form
  POST /rate/<facility_id>/submit/  -> save submission + answers
  GET  /rate/<facility_id>/thanks/  -> thank you page

Admin / coach-manager (in-app):
  GET  /rate/manage/                                -> facility list
  GET  /rate/manage/facility/new/                   -> create
  POST /rate/manage/facility/new/                   -> save
  GET  /rate/manage/facility/<id>/edit/             -> edit facility + questions
  POST /rate/manage/facility/<id>/edit/             -> save
  POST /rate/manage/facility/<id>/delete/           -> delete
  POST /rate/manage/facility/<id>/toggle/           -> toggle is_active
  GET  /rate/manage/dashboard/                      -> all submissions + stats
  GET  /rate/manage/dashboard/<facility_id>/        -> per-facility detail
  GET  /rate/manage/export.xlsx                     -> Excel export (all or filtered)
"""
from datetime import timedelta
from io import BytesIO

from django.contrib import messages
from django.db import transaction
from django.db.models import Avg, Count, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.translation import gettext as _
from django.views.decorators.http import require_POST

from users.mixins import RoleRequiredMixin, get_client_ip
from users.models import Role
from django.views.generic import View, TemplateView

from .models import Answer, Facility, RatingQuestion, Submission


# Anti-spam: one submission per (facility, IP) every this many seconds.
SUBMISSION_COOLDOWN_SECONDS = 60 * 60  # 1 hour


# ─────────────────────────── PUBLIC (no login) ────────────────────────────

def rate_index(request):
    """List all active facilities — landing for the rating system."""
    facilities = Facility.objects.filter(is_active=True).annotate(
        q_count=Count('questions'),
    ).order_by('order', 'name')
    return render(request, 'ratings/index.html', {
        'facilities': facilities,
    })


def rate_facility(request, facility_id):
    """Show the rating form for one facility."""
    facility = get_object_or_404(Facility, id=facility_id, is_active=True)
    questions = list(facility.questions.all())
    if not questions:
        return render(request, 'ratings/no_questions.html', {
            'facility': facility,
        })

    # Has this visitor already rated this facility in the cooldown window?
    if not request.session.session_key:
        request.session.save()
    already = _recent_submission(facility, request)

    return render(request, 'ratings/form.html', {
        'facility': facility,
        'questions': questions,
        'star_range': range(1, 6),
        'already_submitted': already is not None,
    })


@require_POST
def rate_submit(request, facility_id):
    """Process a public rating submission."""
    facility = get_object_or_404(Facility, id=facility_id, is_active=True)
    questions = list(facility.questions.all())
    if not questions:
        messages.error(request, _('This facility has no rating questions yet.'))
        return redirect('ratings:index')

    # Ensure we have a session key for spam dedup
    if not request.session.session_key:
        request.session.save()

    if _recent_submission(facility, request) is not None:
        messages.warning(
            request,
            _('You already submitted a rating for this facility recently. Thank you!'),
        )
        return redirect('ratings:thanks', facility_id=facility.id)

    # Collect answers and validate each question is answered with 1-5
    answers_to_save = []
    errors = {}
    for q in questions:
        raw = request.POST.get(f'q_{q.id}', '').strip()
        try:
            stars = int(raw)
        except (TypeError, ValueError):
            stars = 0
        if stars < 1 or stars > 5:
            errors[q.id] = _('Please pick a star rating from 1 to 5.')
        else:
            answers_to_save.append((q, stars))

    if errors:
        return render(request, 'ratings/form.html', {
            'facility': facility,
            'questions': questions,
            'star_range': range(1, 6),
            'errors': errors,
            'posted': request.POST,
            'already_submitted': False,
        })

    with transaction.atomic():
        sub = Submission.objects.create(
            facility=facility,
            visitor_name=request.POST.get('visitor_name', '').strip()[:120],
            phone=request.POST.get('phone', '').strip()[:30],
            comment=request.POST.get('comment', '').strip(),
            session_key=request.session.session_key or '',
            ip_address=get_client_ip(request),
            user_agent=(request.META.get('HTTP_USER_AGENT', '') or '')[:255],
        )
        Answer.objects.bulk_create([
            Answer(submission=sub, question=q, stars=s)
            for q, s in answers_to_save
        ])

    return redirect('ratings:thanks', facility_id=facility.id)


def rate_thanks(request, facility_id):
    facility = get_object_or_404(Facility, id=facility_id)
    return render(request, 'ratings/thanks.html', {'facility': facility})


def _recent_submission(facility, request):
    """Return the most recent submission by this visitor for this facility,
    or None if outside the cooldown window."""
    cutoff = timezone.now() - timedelta(seconds=SUBMISSION_COOLDOWN_SECONDS)
    session_key = request.session.session_key or ''
    ip = get_client_ip(request)
    qs = Submission.objects.filter(facility=facility, submitted_at__gte=cutoff)
    cond = Q()
    if session_key:
        cond |= Q(session_key=session_key)
    if ip:
        cond |= Q(ip_address=ip)
    if not cond:
        return None
    return qs.filter(cond).order_by('-submitted_at').first()


# ─────────────────────── ADMIN MANAGEMENT (RoleRequired) ───────────────────

ADMIN_ROLES = [Role.ADMIN, Role.COACH_MANAGER]


class ManageFacilityList(RoleRequiredMixin, View):
    required_roles = ADMIN_ROLES

    def get(self, request):
        facilities = Facility.objects.annotate(
            q_count=Count('questions', distinct=True),
            s_count=Count('submissions', distinct=True),
        ).order_by('order', 'name')
        return render(request, 'ratings/manage_list.html', {
            'facilities': facilities,
        })


class ManageFacilityForm(RoleRequiredMixin, View):
    """Create or edit a facility together with its questions (one form)."""
    required_roles = ADMIN_ROLES
    template_name = 'ratings/manage_form.html'

    def _ctx(self, facility, posted=None, errors=None):
        # Build a list of question dicts so the template can render either the
        # saved values or whatever the user just typed in.
        if posted:
            # Reconstruct from POST so user doesn't lose input on error
            qs = []
            for i in range(int(posted.get('q_total', 0) or 0)):
                qs.append({
                    'id': posted.get(f'q_id_{i}', ''),
                    'text': posted.get(f'q_text_{i}', ''),
                    'order': posted.get(f'q_order_{i}', str(i)),
                    'error': (errors or {}).get(f'q_text_{i}'),
                })
            # Pad to at least 5 question rows for editing convenience
            while len(qs) < 5:
                qs.append({'id': '', 'text': '', 'order': str(len(qs)), 'error': None})
        else:
            qs = [
                {'id': str(q.id), 'text': q.text, 'order': str(q.order), 'error': None}
                for q in (facility.questions.all() if facility else [])
            ]
            # Always show at least 5 question slots
            while len(qs) < 5:
                qs.append({'id': '', 'text': '', 'order': str(len(qs)), 'error': None})

        return {
            'facility': facility,
            'questions': qs,
            'posted': posted or {},
            'errors': errors or {},
        }

    def get(self, request, facility_id=None):
        facility = get_object_or_404(Facility, id=facility_id) if facility_id else None
        return render(request, self.template_name, self._ctx(facility))

    def post(self, request, facility_id=None):
        facility = get_object_or_404(Facility, id=facility_id) if facility_id else None
        post = request.POST
        errors = {}

        name = post.get('name', '').strip()
        if not name:
            errors['name'] = _('Name is required.')

        description = post.get('description', '').strip()
        icon = post.get('icon', '⭐').strip()[:8] or '⭐'
        is_active = bool(post.get('is_active'))
        try:
            order = int(post.get('order') or 0)
        except ValueError:
            order = 0

        # Gather questions
        try:
            q_total = int(post.get('q_total') or 0)
        except ValueError:
            q_total = 0
        questions_payload = []
        for i in range(q_total):
            text = post.get(f'q_text_{i}', '').strip()
            qid = post.get(f'q_id_{i}', '').strip()
            try:
                q_order = int(post.get(f'q_order_{i}') or i)
            except ValueError:
                q_order = i
            if not text and not qid:
                continue  # blank row -> ignore
            if not text and qid:
                # Treat blank text on existing question as a deletion
                questions_payload.append({'id': qid, 'delete': True})
                continue
            questions_payload.append({
                'id': qid, 'text': text, 'order': q_order, 'delete': False,
            })

        if not [q for q in questions_payload if not q.get('delete')]:
            errors['non_field'] = _('Add at least one rating question.')

        if errors:
            ctx = self._ctx(facility, posted=post, errors=errors)
            return render(request, self.template_name, ctx)

        with transaction.atomic():
            if facility is None:
                facility = Facility.objects.create(
                    name=name, description=description, icon=icon,
                    is_active=is_active, order=order,
                )
            else:
                facility.name = name
                facility.description = description
                facility.icon = icon
                facility.is_active = is_active
                facility.order = order
                facility.save()

            # Apply question changes
            keep_ids = set()
            for idx, q in enumerate(questions_payload):
                if q.get('delete'):
                    if q['id']:
                        RatingQuestion.objects.filter(
                            id=q['id'], facility=facility,
                        ).delete()
                    continue
                if q['id']:
                    try:
                        obj = RatingQuestion.objects.get(id=q['id'], facility=facility)
                        obj.text = q['text']
                        obj.order = q['order']
                        obj.save()
                        keep_ids.add(obj.id)
                    except RatingQuestion.DoesNotExist:
                        obj = RatingQuestion.objects.create(
                            facility=facility, text=q['text'], order=q['order'],
                        )
                        keep_ids.add(obj.id)
                else:
                    obj = RatingQuestion.objects.create(
                        facility=facility, text=q['text'], order=q['order'],
                    )
                    keep_ids.add(obj.id)

        messages.success(
            request,
            _('Saved “{name}” with {n} question(s).').format(
                name=facility.name,
                n=facility.questions.count(),
            ),
        )
        return redirect('ratings:manage_list')


class ManageFacilityDelete(RoleRequiredMixin, View):
    required_roles = ADMIN_ROLES

    def get(self, request, facility_id):
        facility = get_object_or_404(Facility, id=facility_id)
        return render(request, 'ratings/manage_confirm_delete.html', {
            'facility': facility,
        })

    def post(self, request, facility_id):
        facility = get_object_or_404(Facility, id=facility_id)
        name = facility.name
        facility.delete()
        messages.success(request, _('Deleted “{name}”.').format(name=name))
        return redirect('ratings:manage_list')


@require_POST
def manage_toggle_active(request, facility_id):
    """Quick toggle from the management list."""
    if not (request.user.is_authenticated and (request.user.is_admin or request.user.is_coach_manager)):
        return redirect('login')
    facility = get_object_or_404(Facility, id=facility_id)
    facility.is_active = not facility.is_active
    facility.save(update_fields=['is_active'])
    state = _('opened for rating') if facility.is_active else _('closed')
    messages.success(request, _('“{name}” is now {state}.').format(name=facility.name, state=state))
    return redirect('ratings:manage_list')


class Dashboard(RoleRequiredMixin, View):
    """Overall ratings dashboard — totals, averages, per-facility summary."""
    required_roles = ADMIN_ROLES

    def get(self, request):
        facilities = Facility.objects.annotate(
            s_count=Count('submissions', distinct=True),
        ).order_by('-s_count', 'name')

        facility_rows = []
        for f in facilities:
            avg_overall = Answer.objects.filter(
                submission__facility=f,
            ).aggregate(a=Avg('stars'))['a']
            facility_rows.append({
                'facility': f,
                'submission_count': f.s_count,
                'average': round(avg_overall, 2) if avg_overall else None,
            })

        total_submissions = Submission.objects.count()
        avg_all = Answer.objects.aggregate(a=Avg('stars'))['a']
        latest = Submission.objects.select_related('facility')[:25]

        return render(request, 'ratings/dashboard.html', {
            'facility_rows': facility_rows,
            'total_submissions': total_submissions,
            'total_facilities': facilities.count(),
            'avg_all': round(avg_all, 2) if avg_all else None,
            'latest_submissions': latest,
        })


class FacilityDashboard(RoleRequiredMixin, View):
    """Detail dashboard for a single facility — every submission shown."""
    required_roles = ADMIN_ROLES

    def get(self, request, facility_id):
        facility = get_object_or_404(Facility, id=facility_id)
        questions = list(facility.questions.all())
        # Average per question
        question_rows = []
        for q in questions:
            avg = q.answers.aggregate(a=Avg('stars'))['a']
            question_rows.append({
                'question': q,
                'average': round(avg, 2) if avg else None,
                'count': q.answers.count(),
            })

        submissions = facility.submissions.prefetch_related('answers__question')
        sub_rows = []
        for s in submissions:
            by_q = {a.question_id: a.stars for a in s.answers.all()}
            sub_rows.append({
                'submission': s,
                'answers': [{'q': q, 'stars': by_q.get(q.id)} for q in questions],
            })

        return render(request, 'ratings/dashboard_facility.html', {
            'facility': facility,
            'questions': questions,
            'question_rows': question_rows,
            'submission_rows': sub_rows,
        })


class ExportExcel(RoleRequiredMixin, View):
    """Export all submissions (or one facility's) to Excel."""
    required_roles = ADMIN_ROLES

    def get(self, request):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            messages.error(request, _('Excel export requires the openpyxl package.'))
            return redirect('ratings:dashboard')

        facility_id = request.GET.get('facility')
        if facility_id:
            facilities = Facility.objects.filter(id=facility_id)
        else:
            facilities = Facility.objects.all().order_by('order', 'name')

        wb = Workbook()
        # Remove the default empty sheet — we'll add per-facility sheets
        default_sheet = wb.active
        wb.remove(default_sheet)

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='5A0F0F')  # brand burgundy
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Summary sheet first
        summary = wb.create_sheet(title=str(_('Summary'))[:31])
        summary.append([
            str(_('Facility')), str(_('Submissions')), str(_('Average (1-5)')), str(_('Active?')),
        ])
        for cell in summary[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
        for f in facilities:
            avg = Answer.objects.filter(submission__facility=f).aggregate(a=Avg('stars'))['a']
            summary.append([
                f.name,
                f.submissions.count(),
                round(avg, 2) if avg else None,
                'Yes' if f.is_active else 'No',
            ])
        for col_letter, width in zip('ABCD', [30, 14, 14, 10]):
            summary.column_dimensions[col_letter].width = width

        for facility in facilities:
            questions = list(facility.questions.all())
            sheet = wb.create_sheet(title=facility.name[:31] or 'Facility')

            headers = [
                str(_('Submitted at')),
                str(_('Name')),
                str(_('Phone')),
                str(_('Comment')),
                str(_('IP')),
            ]
            for q in questions:
                headers.append(f'★ {q.text[:80]}')
            headers.append(str(_('Average')))
            sheet.append(headers)
            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center

            submissions = facility.submissions.prefetch_related('answers').order_by('-submitted_at')
            for sub in submissions:
                by_q = {a.question_id: a.stars for a in sub.answers.all()}
                row = [
                    sub.submitted_at.replace(tzinfo=None) if sub.submitted_at else None,
                    sub.visitor_name,
                    sub.phone,
                    sub.comment,
                    sub.ip_address,
                ]
                star_values = [by_q.get(q.id) for q in questions]
                row.extend(star_values)
                non_null = [s for s in star_values if s is not None]
                row.append(round(sum(non_null) / len(non_null), 2) if non_null else None)
                sheet.append(row)

            # Column widths
            sheet.column_dimensions['A'].width = 20
            sheet.column_dimensions['B'].width = 22
            sheet.column_dimensions['C'].width = 18
            sheet.column_dimensions['D'].width = 40
            sheet.column_dimensions['E'].width = 16
            for i in range(len(questions)):
                col_letter = sheet.cell(row=1, column=6 + i).column_letter
                sheet.column_dimensions[col_letter].width = 28
            avg_col_letter = sheet.cell(row=1, column=6 + len(questions)).column_letter
            sheet.column_dimensions[avg_col_letter].width = 12
            sheet.freeze_panes = 'A2'

        # If no facility had submissions, still produce a usable file
        if len(wb.sheetnames) == 1 and summary.max_row == 1:
            summary.append(['—', 0, None, '—'])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        stamp = timezone.localtime().strftime('%Y%m%d-%H%M')
        filename = f'ratings-{stamp}.xlsx'

        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
