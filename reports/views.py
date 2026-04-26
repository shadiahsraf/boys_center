import io
from datetime import date, timedelta
from django.contrib import messages
from django.db.models import Avg, Count
from django.http import HttpResponse
from django.utils.translation import gettext_lazy as _
from django.views.generic import TemplateView, View

from users.mixins import RoleRequiredMixin
from users.models import Role
from attendance.models import AttendanceRecord, AttendanceSession
from evaluations.models import Evaluation


class ReportDashboardView(RoleRequiredMixin, TemplateView):
    required_roles = [Role.ADMIN, Role.COACH_MANAGER, Role.COACH]
    template_name = 'reports/dashboard.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        today = date.today()
        ctx['records_today'] = AttendanceRecord.objects.filter(check_in_time__date=today).count()
        ctx['records_week'] = AttendanceRecord.objects.filter(
            check_in_time__date__gte=today - timedelta(days=7)
        ).count()
        ctx['evals_week'] = Evaluation.objects.filter(
            created_at__date__gte=today - timedelta(days=7)
        ).count()
        return ctx


class AttendanceReportView(RoleRequiredMixin, TemplateView):
    required_roles = [Role.ADMIN, Role.COACH_MANAGER, Role.COACH]
    template_name = 'reports/attendance_report.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        period = self.request.GET.get('period', 'month')
        today = date.today()
        if period == 'week':
            start = today - timedelta(days=7)
        else:
            start = today.replace(day=1)
        sessions = (AttendanceSession.objects
                    .filter(date__gte=start)
                    .select_related('coach')
                    .annotate(record_count=Count('records'))
                    .order_by('-date'))
        ctx.update({'sessions': sessions, 'period': period, 'start': start})
        return ctx


class ExportAttendancePDFView(RoleRequiredMixin, View):
    required_roles = [Role.ADMIN, Role.COACH_MANAGER, Role.COACH]

    def get(self, request, *args, **kwargs):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
        except ImportError:
            return HttpResponse("ReportLab not available", status=500)

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, title="Attendance Report",
                                rightMargin=0.5*inch, leftMargin=0.5*inch,
                                topMargin=0.5*inch, bottomMargin=0.5*inch)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('CustomTitle', parent=styles['Title'],
                                     textColor=colors.HexColor('#0f172a'), fontSize=18)
        elements = [
            Paragraph("Boys Center — Attendance Report", title_style),
            Paragraph(f"Generated: {date.today():%B %d, %Y}", styles['Normal']),
            Spacer(1, 0.2*inch),
        ]

        records = (AttendanceRecord.objects
                   .select_related('user', 'session')
                   .order_by('-check_in_time')[:200])
        data = [['Player', 'Code', 'Session', 'Date', 'Status']]
        for r in records:
            data.append([
                r.user.get_full_name() or r.user.username,
                r.user.member_code or '',
                r.session.title[:30],
                str(r.session.date),
                r.status.upper()
            ])

        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('PADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(t)
        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="attendance_{date.today()}.pdf"'
        return response


class ExportAttendanceExcelView(RoleRequiredMixin, View):
    required_roles = [Role.ADMIN, Role.COACH_MANAGER, Role.COACH]

    def get(self, request, *args, **kwargs):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return HttpResponse("openpyxl not available", status=500)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance"

        # Header styling
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin = Side(border_style='thin', color='cbd5e1')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = ['Player', 'Member Code', 'Session', 'Type', 'Date', 'Coach', 'Status', 'Check-in']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        records = (AttendanceRecord.objects
                   .select_related('user', 'session', 'session__coach')
                   .order_by('-check_in_time'))
        for row_idx, r in enumerate(records, 2):
            ws.cell(row=row_idx, column=1, value=r.user.get_full_name() or r.user.username)
            ws.cell(row=row_idx, column=2, value=r.user.member_code or '')
            ws.cell(row=row_idx, column=3, value=r.session.title)
            ws.cell(row=row_idx, column=4, value=r.session.get_session_type_display())
            ws.cell(row=row_idx, column=5, value=r.session.date.strftime('%Y-%m-%d'))
            ws.cell(row=row_idx, column=6, value=r.session.coach.get_full_name() if r.session.coach else '')
            ws.cell(row=row_idx, column=7, value=r.status)
            ws.cell(row=row_idx, column=8, value=r.check_in_time.strftime('%Y-%m-%d %H:%M'))

        # Column widths
        for col, width in enumerate([24, 14, 28, 12, 12, 24, 12, 18], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        ws.freeze_panes = 'A2'

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="attendance_{date.today()}.xlsx"'
        return response


class ExportEvaluationsExcelView(RoleRequiredMixin, View):
    required_roles = [Role.ADMIN, Role.COACH_MANAGER]

    def get(self, request, *args, **kwargs):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Evaluations"

        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        headers = ['Player', 'Code', 'Coach', 'Sport', 'Performance', 'Behavior', 'Commitment', 'Average', 'Date']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        evals = Evaluation.objects.select_related('coach', 'player').order_by('-created_at')
        for i, e in enumerate(evals, 2):
            ws.cell(row=i, column=1, value=e.player.get_full_name())
            ws.cell(row=i, column=2, value=e.player.member_code or '')
            ws.cell(row=i, column=3, value=e.coach.get_full_name())
            ws.cell(row=i, column=4, value=e.sport)
            ws.cell(row=i, column=5, value=e.performance)
            ws.cell(row=i, column=6, value=e.behavior)
            ws.cell(row=i, column=7, value=e.commitment)
            ws.cell(row=i, column=8, value=float(e.average))
            ws.cell(row=i, column=9, value=e.created_at.strftime('%Y-%m-%d'))

        for col, width in enumerate([24, 14, 24, 16, 14, 14, 14, 12, 14], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        ws.freeze_panes = 'A2'

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="evaluations_{date.today()}.xlsx"'
        return response
